import openpyxl

# 파일 오픈
wb = openpyxl.load_workbook('example.xlsx')

# wb 타입 확인
print(type(wb))

# 현재 워크북에 워크시트 확인
# print(wb.get_sheet_names()) -> 아래 명령어로 대체
print(wb.sheetnames)

# 워크시트 접근
# sheet = wb.get_sheet_by_name('Sheet1') -> 아래 명령어로 대체
sheet = wb['Sheet1']
print(type(sheet))

# 엑티브 시트 사용해 활성화 된 시트 확인
active_sheet = wb.active
print(type(active_sheet))

# 셀에 접근 - 인덱스
cell = sheet['A1']
print(cell)

# 셀에 데이터 가져오기
cell_value = cell.value
print(cell_value)
print('Cell ' + cell.coordinate + ' is ' + cell_value)
print(sheet['A45'].value)

# 셀에 접근 - 키워드 파라미터 사용
keyword_cell = sheet.cell(row=1, column=2)
print(keyword_cell)

# 셀에 데이터 가져오기
keyword_cell_value = sheet.cell(row=1, column=2).value
print(keyword_cell_value)

# for 문 사용 데이터 가져오기
for i in range(1, 8, 2):
	print(i, sheet.cell(row=i, column=2).value)

# 마지막 셀(로우) 가져오기
last_cell_row = sheet.max_row
print(last_cell_row)

# 마지막 셀(컬럼) 가져오기
last_cell_column = sheet.max_column
print(last_cell_column)

# 여러 셀 접근하기
cell_range = sheet['A1' : 'C2']
print(cell_range)
colC = sheet['C']
print(colC)
col_range = sheet['C:D']
print(col_range)
row10 = sheet[10]
print(row10)
row_range = sheet[5:10]
print(row_range)

# 전체 컬럼과 로우에 접근하는 코드
for col_cell in colC:
	print(col_cell.value)

for col in col_range:
	for cell in col:
		print(cell.value)
	print('')

for col_cell in colC:
	print(col_cell.value)

for row in row_range:
	for cell in row:
		print(cell.value)
	print('')