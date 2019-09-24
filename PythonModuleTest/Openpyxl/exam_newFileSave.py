''' 엑셀 문서 만들고 저장하기 '''
from openpyxl import Workbook

from openpyxl import Workbook

wb = Workbook()
print(wb.sheetnames)

# 시트 이름 확인
sheet = wb.active
print(sheet.title)

# 시트이름 변경
sheet.title = '파이썬 업무 자동화'
print(sheet.title)
wb.create_sheet('sheet2')  # 마지막에 추가
wb.create_sheet('sheet3', 1)  # 해당 위치에 추가

# A1 셀의 값 읽기
sheet_a1 = sheet['A1']
print(sheet_a1)

# A1 셀의 값 쓰기
sheet['A1'] = 'hello'

# A1 셀의 값 읽기
sheet_a1_cell = sheet.cell(row=1, column=1)
print(sheet_a1_cell)

# A1 셀의 값 쓰기
sheet.cell(row=1, column=1, value='world!')

# 워크북 저장
wb.save('python.xlsx')
